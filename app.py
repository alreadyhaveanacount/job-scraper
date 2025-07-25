############ IMPORTS ############

from dash import Dash, html, dcc, Input, Output, callback, State
from job_scraper import load_page

import dash_ag_grid as dag
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

from openpyxl import Workbook # Creates excel file
from openpyxl.worksheet.table import Table, TableStyleInfo # For creating and styling tables
from openpyxl.utils.dataframe import dataframe_to_rows # For importing pandas DataSet to excel
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from io import BytesIO

import base64

############ SETTING UP DASH ############

app = Dash()

app.layout = html.Div(children=[
    html.Div(children=[
        dcc.Input(placeholder="Qual cargo você busca?", type="text", id="query"),
        dcc.Input(placeholder="Quantas páginas serão vasculhadas?", type="number", min=1, step=1, max=50, id="page_amount"),
        dcc.Input(placeholder="Pagamento minímo desejado(em $)?", type="number", min=1, id="min_payment"),
        dcc.Dropdown(["full_time", "part_time", "contract"], multi=True, id="wanted_employments", placeholder="Tipos de contrato desejados"),
        dcc.Dropdown(["remote", "hybrid", "on_site"], multi=True, id="workplace_locations", placeholder="Locais de trabalho desejados"),
        dcc.Dropdown(["never", "sometimes", "often"], id="travel_frequency", placeholder="Frequência de viagem"),
        dcc.Dropdown(["USD", "CAD", "EUR", "GBP", "AUD"], value="USD", id="pay_currency", placeholder="Moeda de pagamento"),
        dcc.Dropdown(["entry_level", "mid_level", "senior", "manager", "director", "executive"], id="seniorities", placeholder="Nível mínimo de senioridade"),
        html.Button("Coletar Dados", id="submit", style={"marginTop": "30px"}, n_clicks=0),
    ], id="filters"),
    html.Div(children=[
        html.H2("Dados Coletados", style={"textAlign": "center", "color": "white"}),
        dcc.Loading(
            dag.AgGrid(
                id="scraped_table",
                columnDefs=[],
                rowData=[],
                defaultColDef={
                    "flex": 1,
                    "filter": True,
                    "resizable": True,
                    "sortable": True,
                    "editable": False,
                    "minWidth": 120
                },
                dashGridOptions={
                    "pagination": True,
                    "paginationPageSize": 20,
                    "domLayout": "autoHeight",
                    "animateRows": True,
                    "rowSelection": "single",
                    "enableCellTextSelection": True
                }
            ),
            type="circle",
            color="#00cc96"
        ),
        html.Button("Baixar tabela em Excel", id="btn_excel", style={"marginTop": "30px"}, n_clicks=0),
        dcc.Download(id="download")
    ], id="scraped_stuff"),
])

############ HELPER FUNCTIONS ############

def write_df_to_sheet(dataframe, sheet, t_name, to_format=[], currency_columns={}):
    rows = dataframe_to_rows(dataframe, index=False)

    for r_indx, row in enumerate(rows, 1):
        for c_indx, value in enumerate(row, 1):
            new_cell = sheet.cell(row=r_indx, column=c_indx, value=value)
            column_letter = get_column_letter(c_indx)

            if isinstance(value, str) and sheet.column_dimensions[column_letter].width < len(value) + 2:
                sheet.column_dimensions[column_letter].width = len(value) + 2

            if dataframe.columns[c_indx-1] in currency_columns:
                new_cell.number_format = '$#,###,##0.00'
                sheet.column_dimensions[column_letter].width = max(sheet.column_dimensions[column_letter].width, 13)
    
    t_style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    t = Table(displayName=t_name, ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))

    t.tableStyleInfo = t_style

    sheet.add_table(t)

    if to_format is not None:
        for column in to_format:
            column_letter = get_column_letter(dataframe.columns.get_loc(column) + 1)
            format_range = f"{column_letter}1:{column_letter}{sheet.max_row}"
            format_rule = ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_color="FFEB84", mid_value=50, end_type="max", end_color="63BE7B")
            sheet.conditional_formatting.add(format_range, format_rule)

############ CALLBACKS ############

@callback(
    [
        Output(component_id="scraped_table", component_property="columnDefs"),
        Output(component_id="scraped_table", component_property="rowData")
    ],
    Input(component_id="submit", component_property="n_clicks"),
    [
        State(component_id="query", component_property="value"),
        State(component_id="wanted_employments", component_property="value"),
        State(component_id="workplace_locations", component_property="value"),
        State(component_id="seniorities", component_property="value"),
        State(component_id="travel_frequency", component_property="value"),
        State(component_id="min_payment", component_property="value"),
        State(component_id="pay_currency", component_property="value"),
        State(component_id="page_amount", component_property="value")
    ]
, prevent_initial_call=True)
def scrape_data(btn_clicks, query, employment_types, workplace_location, seniority, travel_frequency, compensation, compensation_currency, page_amount):
    buffer = []

    arguments = {
        "compensationCurrency": compensation_currency
    }

    if page_amount is None: page_amount = 0
    if query is not None and query != "": arguments["query"] = query
    if employment_types is not None and len(employment_types) > 0: arguments["employmentType"] = employment_types
    if workplace_location is not None and len(workplace_location) > 0: arguments["workplaceLocation"] = workplace_location
    if seniority is not None and len(seniority) > 0: arguments["seniority"] = seniority
    if travel_frequency is not None and len(travel_frequency) > 0: arguments["travelFrequency"] = travel_frequency
    if compensation is not None and compensation > 0: arguments["compensation"] = compensation * 100

    page_amount = round(max(min(page_amount, 50), 1))

    for current_page in range(0, page_amount+1):
        if current_page > 1:
            arguments["page"] = current_page
        
        prev_len = len(buffer)
        load_page(buffer, arguments)

        if prev_len == len(buffer):
            break

    scraped_df = pd.DataFrame(buffer)

    money_columns = {"Minimal Salary", "Maximum Salary", "Average Hourly Salary"}

    return [{"field": c} if c not in money_columns else {"field": c, "valueFormatter": {"function": "d3.format('($,.2f')(params.value)"}} for c in scraped_df.columns], scraped_df.to_dict(orient="records")

@callback(
    Output(component_id="download", component_property="data"),
    Input(component_id="btn_excel", component_property="n_clicks"),
    State(component_id="scraped_table", component_property="rowData")
, prevent_initial_call=True)
def download_excel(clicks, df_data):
    df = pd.DataFrame(df_data)

    excel_workbook = Workbook()
    excel_workbook.active.title = "Scraped Data"
    write_df_to_sheet(df, excel_workbook.active, "scraped_table", ["Average Hourly Salary"], {"Minimal Salary", "Maximum Salary", "Average Hourly Salary"})

    excel_stream = BytesIO()

    excel_workbook.save(excel_stream)

    excel_stream.seek(0)

    b64_url = base64.b64encode(excel_stream.read()).decode()

    return dict(
        content=b64_url, 
        type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="scraped.xlsx",
        base64=True
    )

############ STARTING DASHBOARD ############

if __name__ == '__main__':
    app.run(debug=False)